import math
from typing import Any, Dict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def format_value(val: Any) -> str:
    """Format a number (real or complex) in scientific notation.

    Returns:
        str: For complex -> 'a.bce±dd + d.efeg±hhj', else 'a.bce±dd'.
    """
    if val is None:
        return "nan"

    # Handle numpy scalars
    if isinstance(val, (np.generic,)):
        val = np.asarray(val).item()

    # Complex numbers
    if isinstance(val, complex) or np.iscomplex(val):
        real = float(np.real(val))
        imag = float(np.imag(val))
        sign = "+" if imag >= 0 else "-"
        return f"{real:.2e} {sign} {abs(imag):.2e}j"

    # Non-complex numeric
    try:
        fval = float(val)
        if math.isnan(fval):
            return "nan"
        if math.isinf(fval):
            return "inf" if fval > 0 else "-inf"
        return f"{fval:.2e}"
    except (TypeError, ValueError):
        # Fallback: stringify anything else
        return str(val)


def format_freq(val: Any) -> str:
    """Convert a (possibly complex) number to frequency in Hz as scientific notation.

    Uses |val|/(2π). Non-numerics -> 'nan'.
    """
    try:
        magnitude = abs(complex(val))
        hz = magnitude / (2 * np.pi)
        return f"{hz:.2e}"
    except Exception:
        return "nan"


def save_df(dfs: Dict[str, pd.DataFrame], filename_noext: str) -> None:
    """
    Save multiple pandas DataFrames to a single Excel file with formatted layout.

    Parameters:
        dfs (dict): Mapping {sheet_title: DataFrame}. Titles are written as section headers
                    in a single worksheet called 'All_DataFrames'.
        filename_noext (str): Output filename without extension.

    Notes:
        - Writes each DataFrame one after another with a blank row in between.
        - Applies borders, header fill, centered alignment, and auto column widths.
    """
    if not dfs:
        print("No DataFrames provided.")
        return

    filename = f"{filename_noext}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "All_DataFrames"

    # Predefined styles (reused objects are cheaper than creating per-cell)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    title_font = Font(bold=True, size=12, color="1F4E78")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    current_row = 1
    # Track max width per column index across the whole sheet to avoid ws.columns pass
    max_widths: Dict[int, int] = {}

    for name, df in dfs.items():
        # Defensive: ensure DataFrame
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame(df)

        # Section title (single cell)
        title_cell = ws.cell(row=current_row, column=1, value=str(name))
        title_cell.font = title_font
        current_row += 1

        # Write DataFrame rows (header + data)
        # dataframe_to_rows yields lists for header row then data rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            row_number = current_row + r_idx
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_number, column=c_idx, value=value)
                cell.border = border
                cell.alignment = center_align

                if r_idx == 0:  # header row
                    cell.fill = header_fill
                    cell.font = header_font

                # Track width (string length)
                val_len = len(str(value)) if value is not None else 0
                if val_len > max_widths.get(c_idx, 0):
                    max_widths[c_idx] = val_len

            # After finishing all columns for this row, continue
        # Move cursor past the written DF (header + data rows)
        current_row += (len(df) + 1)

        # Blank spacer row
        current_row += 1

    # Auto-adjust column widths once, based on collected max lengths
    for c_idx, width in max_widths.items():
        col_letter = get_column_letter(c_idx)
        # +2 padding for readability, cap a bit for extreme cases
        ws.column_dimensions[col_letter].width = min(width + 2, 80)

    wb.save(filename)
    print(f"Output file '{filename}' saved in the current folder.")
